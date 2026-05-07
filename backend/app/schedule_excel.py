"""Load master schedule rows from `Schedules.xlsx` (Skipper Pools format).

Maps spreadsheet columns to internal `task_key` values used by the app.
See `constants.TASK_DEFINITIONS` for the canonical 33-step checklist.

Row 1 headers must match `_SCHEDULE_SHEET_HEADERS` (same order as tasks).
"""
from __future__ import annotations

import logging
import warnings
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .config import PROJECT_ROOT, settings
from .constants import (
    STATUS_COMPLETED,
    STATUS_IN_PROGRESS,
    STATUS_NOT_STARTED,
    TASK_DEFINITIONS,
)
from .models import Job, JobTask

logger = logging.getLogger("skipper.schedule_excel")


def default_excel_path() -> Path:
    """Resolved path to Schedules.xlsx (env `SCHEDULE_XLSX_PATH` / settings)."""
    if settings.schedule_xlsx_path:
        return Path(settings.schedule_xlsx_path).expanduser().resolve()
    return (PROJECT_ROOT / "Schedules.xlsx").resolve()


def _utc(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _cell_to_state(cell: Any) -> Optional[Dict[str, Any]]:
    """Turn one Excel cell into task fields, or None if empty."""
    if cell is None:
        return None
    if isinstance(cell, datetime):
        d = _utc(cell)
        return {
            "status": STATUS_COMPLETED,
            "value": d.date().isoformat(),
            "completed_at": d,
            "note": None,
        }
    # openpyxl may return date objects for date-only cells in some cases
    if isinstance(cell, date) and not isinstance(cell, datetime):
        dt = datetime(cell.year, cell.month, cell.day, tzinfo=timezone.utc)
        return {
            "status": STATUS_COMPLETED,
            "value": cell.isoformat(),
            "completed_at": dt,
            "note": None,
        }
    s = str(cell).strip()
    if not s:
        return None
    return {
        "status": STATUS_IN_PROGRESS,
        "value": s[:255],
        "completed_at": None,
        "note": s if len(s) > 255 else None,
    }


def _build_header_map(header_row: Sequence[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip()
        if key:
            m[key] = i
    return m


# First-row strings in Schedules.xlsx, in the same order as TASK_DEFINITIONS.
# Step 5 uses a longer column title in Excel while the app label is "QI 1".
_SCHEDULE_SHEET_HEADERS: Tuple[str, ...] = (
    "Permit Application",
    "Permit Received",
    "Excavation",
    "Form & Steel",
    "Quality Inspection (QI) 1",
    "Shell Bonding Inspection",
    "Shell Steel Inspection",
    "Rough Plumbing Inspection",
    "Gunite",
    "QI 2",
    "Survey",
    "Backfill",
    "Plumbing",
    "QI 3",
    "Electrical Inspection",
    "Coping",
    "Tile",
    "QI 4",
    "Rail Anchor/Grid Inspection",
    "Inspection",
    "Water/Sub-Deck Installation",
    "Paver Installation",
    "QI 5",
    "Equipment Installation",
    "Equipment Wiring",
    "QI 6",
    "Screen/Fence",
    "Electric Inspection",
    "Safety Inspection",
    "Plaster",
    "Startup",
    "Final QI",
    "Final Inspection",
)

if len(_SCHEDULE_SHEET_HEADERS) != len(TASK_DEFINITIONS):
    raise RuntimeError(
        "_SCHEDULE_SHEET_HEADERS and TASK_DEFINITIONS must have the same length"
    )


def _row_task_states(row: Sequence[Any], hdr: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    states: Dict[str, Dict[str, Any]] = {}
    for excel_header, (task_key, _label) in zip(_SCHEDULE_SHEET_HEADERS, TASK_DEFINITIONS):
        idx = hdr.get(excel_header)
        if idx is None:
            continue
        st = _cell_to_state(row[idx] if idx < len(row) else None)
        if st:
            states[task_key] = st
    return states


def build_job_tasks_from_states(states: Dict[str, Dict[str, Any]]) -> List[JobTask]:
    """Create JobTask rows for one job from a task_key -> partial state map."""
    tasks: List[JobTask] = []
    for index, (key, label) in enumerate(TASK_DEFINITIONS):
        spec = states.get(key, {})
        status = spec.get("status", STATUS_NOT_STARTED)
        tasks.append(
            JobTask(
                task_key=key,
                task_label=label,
                status=status,
                value=spec.get("value"),
                note=spec.get("note"),
                completed_at=spec.get("completed_at"),
                sort_order=index,
            )
        )
    return tasks


def load_jobs_from_excel(path: Optional[Path] = None) -> List[Job]:
    """Parse the Schedules sheet into ORM Job objects (not persisted)."""
    xlsx = path or default_excel_path()
    if not xlsx.is_file():
        raise FileNotFoundError(f"Excel file not found: {xlsx}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        from openpyxl import load_workbook  # lazy import

        wb = load_workbook(str(xlsx), read_only=True, data_only=True)
        try:
            names = wb.sheetnames
            if "Schedules" not in names:
                raise ValueError(
                    f"Workbook has no 'Schedules' sheet. Found: {names!r}"
                )
            ws = wb["Schedules"]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

    if not rows:
        return []

    hdr = _build_header_map(rows[0])

    # Required columns
    for required in ("Job Name", "Address"):
        if required not in hdr:
            raise ValueError(f"Schedules sheet missing required column: {required!r}")

    jobs: List[Job] = []
    for raw in rows[1:]:
        if not raw:
            continue
        name = raw[hdr["Job Name"]]
        if name is None or not str(name).strip():
            continue

        pool = _safe_cell(raw, hdr.get("Pool (P) or Pool Spa (PS)"))
        permit = _safe_cell(raw, hdr.get("Permit Number"))
        mgr = _safe_cell(raw, hdr.get("Field Manager"))
        row_num = raw[0] if raw else None

        notes_parts: List[str] = []
        if row_num is not None and str(row_num).strip() != "":
            try:
                notes_parts.append(f"Master schedule row: {int(row_num)}")
            except (TypeError, ValueError):
                notes_parts.append(f"Master schedule row: {row_num}")

        states = _row_task_states(raw, hdr)
        tasks = build_job_tasks_from_states(states)

        jobs.append(
            Job(
                customer_name=str(name).strip(),
                address=str(_safe_cell(raw, hdr["Address"]) or "").strip() or None,
                pool_type=str(pool).strip() if pool else None,
                permit_number=str(permit).strip() if permit else None,
                field_manager=str(mgr).strip() if mgr else None,
                notes="\n".join(notes_parts) if notes_parts else None,
                tasks=tasks,
            )
        )

    logger.info("Loaded %d job(s) from %s", len(jobs), xlsx)
    return jobs


def _safe_cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]
